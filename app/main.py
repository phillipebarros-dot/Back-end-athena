"""
FastAPI Application — Backend da Athena v3.

Substitui os 7 webhooks do n8n + todo o fluxo de orquestração.
Deploy: Cloud Run (mesma infra dos MCPs do Camilo).

Endpoints equivalentes aos webhooks n8n:
  POST /chat              ← When chat message received
  POST /conversations     ← Webhook Conversas (list/create/updateTitle)
  POST /history           ← Webhook Historico
  POST /save-message      ← Webhook Salvar Mensagem
  POST /compact           ← Webhook Compactacao
  POST /feedback          ← Webhook Feedback
  POST /audit             ← Webhook Auditoria
  POST /tts               ← Webhook TTS
  POST /export            ← Webhook Export Sheets
  GET  /health            ← Health check (novo)
"""

from __future__ import annotations

import base64
import logging
import hmac
import os
import time

import pydantic
from google.cloud import bigquery
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import traceback

from app.config import settings
from app.models import (
    AuditRequest,
    ChatRequest,
    ChatResponse,
    CompactRequest,
    ConversationAction,
    ConversationRequest,
    ExportRequest,
    FeedbackRequest,
    HistoryRequest,
    ProvenanceSource,
    SaveMessageRequest,
    TTSRequest,
    TTSResponse,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# ============================================================================
# Lifespan — inicialização de recursos no startup (FIX C2)
# ============================================================================

from contextlib import asynccontextmanager


@asynccontextmanager
async def lifespan(app_instance: FastAPI):
    """Startup: inicializa checkpointer Postgres (cria tabelas se necessário).

    FIX C2: AsyncPostgresSaver precisa de setup() para criar as tabelas
    de checkpoint. Sem isso, a primeira leitura/escrita estoura UndefinedTable.
    """
    from app.agent.graph import initialize_checkpointer
    await initialize_checkpointer()
    logger.info("Lifespan startup completo.")
    yield
    logger.info("Lifespan shutdown.")


# ============================================================================
# App
# ============================================================================

app = FastAPI(
    title="Athena Backend",
    description="Backend Python da Athena — agente LLM de mídia e planejamento",
    version="3.0.0",
    docs_url="/docs" if settings.debug else None,  # Swagger só em dev
    redoc_url=None,
    lifespan=lifespan,  # FIX C2: inicializa checkpointer no startup
)

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"Global exception handler caught: {exc}", exc_info=True)
    return JSONResponse(
        status_code=500,
        content={"detail": str(exc), "type": type(exc).__name__, "trace": traceback.format_exc()}
    )


# CORS — restrito ao frontend da Athena (configura via env var)
_cors_raw = os.getenv("CORS_ALLOWED_ORIGINS", "")
CORS_ORIGINS = [o.strip() for o in _cors_raw.split(",") if o.strip()] if _cors_raw else []

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["POST", "GET", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "x-api-key"],
)


# ============================================================================
# Auth Middleware — equivalente ao httpHeaderAuth do n8n
# ============================================================================

AUTH_TOKEN = settings.mcp.auth_token  # Reutiliza o mesmo token do MCP

from collections import defaultdict
RATE_LIMIT = 100
RATE_LIMIT_WINDOW = 60
rate_limit_records = defaultdict(list)

@app.middleware("http")
async def rate_limit_middleware(request: Request, call_next):
    """Limita a taxa de requisições por IP."""
    if request.url.path in {"/health", "/docs", "/openapi.json"}:
        return await call_next(request)
        
    client_ip = request.client.host if request.client else "unknown"
    now = time.time()
    
    rate_limit_records[client_ip] = [t for t in rate_limit_records[client_ip] if now - t < RATE_LIMIT_WINDOW]
    
    if len(rate_limit_records[client_ip]) >= RATE_LIMIT:
        return JSONResponse(status_code=429, content={"detail": "Too Many Requests"})
        
    rate_limit_records[client_ip].append(now)
    return await call_next(request)


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    """Valida header de autenticação.

    Equivale ao httpHeaderAuth do n8n (Webhook Entrada Chat).
    Endpoints /health e /docs são públicos.
    """
    public_paths = {"/health", "/docs", "/openapi.json"}
    if request.url.path in public_paths:
        return await call_next(request)

    # Em modo debug, aceita qualquer request
    if settings.debug:
        return await call_next(request)

    # Verifica Bearer token ou x-api-key
    auth_header = request.headers.get("authorization", "")
    api_key = request.headers.get("x-api-key", "")
    token = auth_header.replace("Bearer ", "") if auth_header.startswith("Bearer ") else api_key

    if not token or not hmac.compare_digest(token.strip(), AUTH_TOKEN.strip()):
        return JSONResponse(status_code=401, content={"detail": "Unauthorized"})

    return await call_next(request)


# ============================================================================
# Health Check
# ============================================================================

@app.get("/health")
async def health():
    """Health check para Cloud Run."""
    return {"status": "ok", "version": "3.0.0"}


# ============================================================================
# POST /chat — Substitui o fluxo principal do n8n
# (chatTrigger → Sanitizador → Agent → Validador → TTS → Resposta → Log)
# ============================================================================

@app.post("/chat", response_model=ChatResponse)
async def chat(request: ChatRequest):
    """Endpoint principal de chat.

    O Pydantic (ChatRequest) já sanitiza a entrada — substitui o node
    "Sanitizador de Entrada" do n8n.
    """
    start_time = time.time()

    try:
        # Import lazy para não travar o startup
        from app.agent.graph import get_agent
        from app.agent.tools import get_all_tools
        from app.services.response_validator import validate_response

        tools = get_all_tools()
        agent = await get_agent(tools=tools, cliente=request.client)

        # Config do thread — identifica a conversa para o checkpointer
        config = {"configurable": {"thread_id": request.conversation_id}}

        # Invoca o agente
        result = await agent.ainvoke(
            {"messages": [("user", request.message)]},
            config=config,
        )

        # Extrai a resposta do agente (última mensagem do assistant)
        raw_output = result["messages"][-1].content

        # Validador de Resposta — equivalente ao node do n8n
        validated = validate_response(raw_output)
        output = validated["output"]
        attachment = validated.get("attachment")

        # ── Proveniência: extrai queries, tabelas e fontes das tool calls ──
        sources, queries, tables = [], [], []
        for m in result["messages"]:
            # SQL nas chamadas de tool (BigQuery / MCP)
            for tc in (getattr(m, "tool_calls", None) or []):
                args = tc.get("args", {}) if isinstance(tc, dict) else getattr(tc, "args", {})
                sql = args.get("sql") or args.get("query")
                if sql:
                    queries.append(sql)
                tbl = args.get("table") or args.get("dataset")
                if tbl:
                    tables.append(tbl)
            # Nome da tool executada vira fonte
            name = getattr(m, "name", None)
            if name:
                sources.append(name)

        latency_ms = int((time.time() - start_time) * 1000)
        logger.info(
            "Chat processado: user=%s, conv=%s, client=%s, latency=%dms, sources=%s",
            request.user_id,
            request.conversation_id,
            request.client,
            latency_ms,
            sources,
        )

        # Registrador de Requisicao — equivalente ao node BQ INSERT do n8n
        try:
            from app.services.bq_service import get_bq_service
            bq = get_bq_service()
            bq.insert_log(
                session_id=request.conversation_id,
                user_input=request.message,
                agent_response=output,
                latency_ms=latency_ms,
            )
        except Exception as log_err:
            logger.warning("Falha ao registrar log: %s", log_err)

        # TTS inline — Google Cloud TTS (pt-BR nativo) com fallback OpenAI
        audio_b64 = None
        if request.is_audio:
            tts_text = output[:5000]  # Limita texto para TTS

            # Tentar Google Cloud TTS primeiro (voz pt-BR nativa)
            if settings.tts.provider == "google":
                try:
                    from google.cloud import texttospeech
                    tts_client = texttospeech.TextToSpeechClient()
                    synthesis_input = texttospeech.SynthesisInput(text=tts_text)
                    voice_params = texttospeech.VoiceSelectionParams(
                        language_code=settings.tts.google_language,
                        name=settings.tts.google_voice,
                        ssml_gender=texttospeech.SsmlVoiceGender.MALE,
                    )
                    audio_config = texttospeech.AudioConfig(
                        audio_encoding=texttospeech.AudioEncoding.MP3,
                        speaking_rate=settings.tts.google_speaking_rate,
                        pitch=0.0,
                    )
                    tts_response = tts_client.synthesize_speech(
                        input=synthesis_input,
                        voice=voice_params,
                        audio_config=audio_config,
                    )
                    audio_b64 = base64.b64encode(tts_response.audio_content).decode("utf-8")
                    logger.info("TTS Google Cloud (pt-BR-Neural2-B) gerado com sucesso")
                except Exception as google_tts_err:
                    logger.warning("Google TTS falhou, tentando OpenAI: %s", google_tts_err)

            # Fallback OpenAI ou provider openai
            if audio_b64 is None and settings.tts.openai_api_key:
                try:
                    import openai
                    tts_client = openai.OpenAI(api_key=settings.tts.openai_api_key)
                    tts_response = tts_client.audio.speech.create(
                        model=settings.tts.model,
                        voice=settings.tts.voice,
                        input=tts_text,
                    )
                    audio_b64 = base64.b64encode(tts_response.content).decode("utf-8")
                    logger.info("TTS OpenAI (%s) gerado com sucesso", settings.tts.voice)
                except Exception as tts_err:
                    logger.warning("Falha no TTS OpenAI: %s", tts_err)

        return ChatResponse(
            output=output,
            conversation_id=request.conversation_id,
            latency_ms=latency_ms,
            attachment=attachment,
            audio=audio_b64,
            query="\n\n".join(dict.fromkeys(queries)) or None,
            tables=list(dict.fromkeys(tables)) or None,
            sources=[ProvenanceSource(label=s) for s in dict.fromkeys(sources)] or None,
        )

    except Exception as e:
        logger.error("Erro no chat: %s", e, exc_info=True)
        latency_ms = int((time.time() - start_time) * 1000)
        # FIX C4: em debug, mostra o detalhe do erro em vez de mensagem genérica
        error_detail = (
            f"Erro interno ({type(e).__name__}): {e}"
            if settings.debug
            else "Desculpe, ocorreu um erro ao processar sua consulta. Tente novamente."
        )
        return ChatResponse(
            output=error_detail,
            conversation_id=request.conversation_id,
            latency_ms=latency_ms,
        )


# ============================================================================
# POST /conversations — Substitui o Webhook Conversas do n8n (14 nodes)
# ============================================================================

@app.post("/conversations")
async def conversations(request: ConversationRequest):
    """CRUD de conversas.

    Substitui: Webhook Conversas → Sanitizar → Acao Listar? → BQ Listar →
    Formatar → Resposta (e variantes Create/UpdateTitle).
    """
    from app.services.bq_service import get_bq_service
    bq = get_bq_service()

    if request.action == ConversationAction.LIST:
        convs = bq.list_conversations(request.user_id)
        return {"conversations": convs}

    elif request.action == ConversationAction.CREATE:
        if not request.conversation_id:
            raise HTTPException(status_code=400, detail="conversation_id é obrigatório para criação.")
        result = bq.create_conversation(
            user_id=request.user_id,
            conversation_id=request.conversation_id,
            title=request.title or "Nova conversa",
        )
        return {"success": True, "action": "create", **result}

    elif request.action == ConversationAction.UPDATE_TITLE:
        if not request.conversation_id or not request.title:
            raise HTTPException(status_code=400, detail="conversation_id e title são obrigatórios.")
        bq.update_title(request.conversation_id, request.title)
        return {"success": True}

    elif request.action == ConversationAction.DELETE:
        if not request.conversation_id:
            raise HTTPException(status_code=400, detail="conversation_id é obrigatório para deleção.")
        bq.delete_conversation(request.conversation_id)
        return {"success": True, "action": "delete"}


# ============================================================================
# POST /history — Substitui o Webhook Historico do n8n (5 nodes)
# ============================================================================

@app.post("/history")
async def history(request: HistoryRequest):
    """Carrega histórico de mensagens de uma conversa.

    Substitui: Webhook Historico → Sanitizar → BQ Carregar → Formatar → Resposta.
    """
    from app.services.bq_service import get_bq_service
    bq = get_bq_service()
    messages = bq.load_history(request.conversation_id, request.limit)
    return {"messages": messages}


# ============================================================================
# POST /save-message — Substitui o Webhook Salvar Mensagem do n8n (5 nodes)
# ============================================================================

@app.post("/save-message")
async def save_message(request: SaveMessageRequest):
    """Salva uma mensagem individual.

    Substitui: Webhook Salvar → Sanitizar → BQ Inserir → BQ Atualizar Contagem → Resposta.
    """
    from app.services.bq_service import get_bq_service
    bq = get_bq_service()
    result = bq.save_message(
        conversation_id=request.conversation_id,
        user_id=request.user_id,
        role=request.role,
        content=request.content,
    )
    return {"success": True, **result}


# ============================================================================
# POST /compact — Substitui o Webhook Compactacao do n8n (12 nodes)
# ============================================================================

@app.post("/compact")
async def compact(request: CompactRequest):
    """Compacta mensagens antigas de uma conversa.

    Substitui: Webhook Compactacao → Config → BQ Contar → Verificar Limite →
    BQ Carregar Antigas → Montar Transcricao → Sumarizar (Haiku) →
    BQ Salvar Resumo → BQ Marcar Compactadas → Resposta.
    """
    from app.services.bq_service import get_bq_service
    bq = get_bq_service()

    # Verificar limite
    count = bq.count_active_messages(request.conversation_id)
    if count <= request.threshold:
        return {"compacted": False, "reason": f"Apenas {count} mensagens ativas (limite: {request.threshold})."}

    # Carregar mensagens antigas
    old_msgs = bq.load_old_messages(request.conversation_id, request.keep_recent)
    if not old_msgs:
        return {"compacted": False, "reason": "Nenhuma mensagem antiga para compactar."}

    # Montar transcrição
    transcription = "\n".join(
        f"[{m['role']}]: {m['content']}" for m in old_msgs
    )

    # Sumarizar com Haiku
    try:
        from app.agent.graph import _build_summarizer_llm
        haiku = _build_summarizer_llm()
        summary_result = haiku.invoke(
            f"Resuma a seguinte conversa em português, mantendo os pontos principais, "
            f"dados mencionados e decisões tomadas. Seja conciso mas completo:\n\n{transcription}"
        )
        summary = summary_result.content
    except Exception as e:
        logger.error("Erro na sumarização: %s", e)
        return {"compacted": False, "reason": f"Erro na sumarização: {e}"}

    # Salvar resumo e marcar compactadas
    bq.save_summary(request.conversation_id, "system", summary)
    bq.mark_compacted([m["message_id"] for m in old_msgs])

    return {
        "compacted": True,
        "messages_compacted": len(old_msgs),
        "summary_length": len(summary),
    }


# ============================================================================
# POST /feedback — Substitui o Webhook Feedback do n8n (4 nodes)
# ============================================================================

@app.post("/feedback")
async def feedback(request: FeedbackRequest):
    """Recebe feedback positivo/negativo por mensagem.

    Substitui: Webhook Feedback → Sanitizar → BQ Inserir → Resposta.
    """
    from app.services.bq_service import get_bq_service
    bq = get_bq_service()
    bq.save_feedback(
        user_id=request.user_id,
        message_id=request.message_id,
        rating=request.rating.value,
        conversation_id=request.conversation_id or "",
        user_query=request.user_query,
        assistant_response=request.assistant_response,
        comment=request.comment,
    )
    return {"success": True}


# ============================================================================
# Audit helper functions
# ============================================================================

def _get_system_stats(bq) -> dict:
    """Métricas de sistema: latência, custo estimado, taxa sem resultado, tokens."""
    try:
        sql = f"""
            WITH stats AS (
                SELECT
                    COUNT(*) as total_msgs,
                    COUNTIF(role = 'assistant' AND (content IS NULL OR TRIM(content) = '' OR content LIKE '%não encontr%' OR content LIKE '%não conseg%')) as no_result_count,
                    AVG(TIMESTAMP_DIFF(
                        LEAD(timestamp) OVER (PARTITION BY conversation_id ORDER BY timestamp),
                        timestamp, SECOND
                    )) as avg_latency_sec
                FROM `{settings.bq.project_persistence}.{settings.bq.dataset_persistence}.athena_messages`
                WHERE timestamp >= TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL 30 DAY)
                    AND role IN ('user', 'assistant')
            )
            SELECT
                total_msgs,
                no_result_count,
                SAFE_DIVIDE(no_result_count, NULLIF(total_msgs, 0)) * 100 as no_result_pct,
                COALESCE(avg_latency_sec, 0) as avg_latency_sec
            FROM stats
        """
        result = bq._client_persistence.query(sql, timeout=15)
        rows = [dict(r) for r in result.result(timeout=15)]
        row = rows[0] if rows else {}
        return {
            "total_messages_30d": int(row.get("total_msgs", 0)),
            "no_result_count": int(row.get("no_result_count", 0)),
            "no_result_pct": round(float(row.get("no_result_pct", 0)), 1),
            "avg_latency_sec": round(float(row.get("avg_latency_sec", 0)), 1),
            "estimated_cost_month_usd": round(float(row.get("total_msgs", 0)) * 0.008, 2),
            "model": os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-20250514"),
        }
    except Exception as e:
        logger.warning("Falha ao obter system_stats: %s", e)
        return {"error": str(e)}


def _get_mcp_health() -> dict:
    """Verifica saúde dos servidores MCP configurados."""
    import httpx
    mcp_servers = [
        ("publi_consulta", settings.mcp.publi_url),
        ("pesquisas_kantar", settings.mcp.pesquisas_url),
        ("exportar_sheets", settings.mcp.export_url),
        ("digital", settings.mcp.midia_online_url),
    ]
    results = []
    for name, url in mcp_servers:
        if not url or not url.startswith("http"):
            results.append({"name": name, "status": "not_configured", "code": 0})
            continue
        try:
            # Strip trailing /mcp path only (keep hostname intact)
            if url.endswith("/mcp"):
                health_url = url[:-4] + "/health"
            else:
                health_url = url.rstrip("/") + "/health"
            headers = {}
            if settings.mcp.auth_token:
                headers["Authorization"] = f"Bearer {settings.mcp.auth_token}"
            logging.info(f"MCP health check: {name} -> {health_url}")
            r = httpx.get(health_url, timeout=8, headers=headers)
            if r.status_code == 200:
                results.append({"name": name, "status": "ok", "code": 200})
            elif r.status_code == 404:
                # Sem /health -> assume ok se não for 5xx
                base = health_url.rsplit("/health", 1)[0]
                r2 = httpx.get(base, timeout=5, headers=headers)
                results.append({"name": name, "status": "ok" if r2.status_code < 500 else "error", "code": r2.status_code})
            else:
                results.append({"name": name, "status": "error", "code": r.status_code})
        except Exception as e:
            results.append({"name": name, "status": "unreachable", "error": str(e)})
    return {"servers": results}


# ============================================================================
# POST /audit — Substitui o Webhook Auditoria do n8n (5 nodes)
# ============================================================================

@app.post("/audit")
async def audit(request: AuditRequest):
    """Métricas para dashboard admin.

    Substitui: Webhook Auditoria → Validar Admin → Roteador → BQ Query → Formatar → Resposta.
    """
    # Validação admin server-side — consulta athena_users.role no BigQuery
    # Fallback: se BQ falhar, usa lista hardcoded da env var
    from app.services.bq_service import get_bq_service as _get_bq
    try:
        _is_admin = _get_bq().is_admin(request.user_email)
    except Exception:
        _is_admin = request.user_email in settings.security.admin_emails
    if not _is_admin:
        raise HTTPException(status_code=403, detail="Acesso administrativo negado.")

    from app.services.bq_service import get_bq_service
    bq = get_bq_service()

    query = request.query

    if query == "kpis":
        data = bq.audit_kpis()
    elif query == "recent_activity":
        data = bq.audit_recent_activity()
    elif query == "recent_feedback":
        data = bq.audit_recent_feedback()
    elif query == "top_users":
        data = bq.audit_top_users()
    elif query == "all_conversations":
        data = bq.audit_all_conversations(
            date_from=request.date_from,
            date_to=request.date_to,
        )
    elif query == "conversation_messages":
        if not request.conversation_id:
            raise HTTPException(status_code=400, detail="conversation_id obrigatório.")
        data = bq.audit_conversation_messages(request.conversation_id)
    elif query == "system_stats":
        # Admin: latência, custo, sem resultado, tokens
        data = _get_system_stats(bq)
    elif query == "mcp_health":
        # Admin: saúde dos servidores MCP
        data = _get_mcp_health()
    else:
        data = {"message": f"Query '{query}' não suportada. Use: kpis, recent_activity, recent_feedback, top_users, all_conversations, conversation_messages, system_stats, mcp_health."}

    return {"query": query, "data": data}


# ============================================================================
# POST /users — Gerenciamento de usuários (RBAC)
# ============================================================================

@app.post("/users")
async def users(request: Request):
    """CRUD de usuários — athena_users no BigQuery.

    Actions:
      - list: lista todos os usuários
      - upsert: cria/atualiza usuário (chamado no login OAuth)
      - update_role: muda role de um usuário (só admin)
      - check: verifica se email é admin
    """
    from app.services.bq_service import get_bq_service
    bq = get_bq_service()

    body = await request.json()
    action = body.get("action", "")

    if action == "list":
        # Só admin pode listar usuários
        caller_email = body.get("user_email", "")
        try:
            caller_is_admin = bq.is_admin(caller_email)
        except Exception:
            caller_is_admin = caller_email in settings.security.admin_emails
        if not caller_is_admin:
            raise HTTPException(status_code=403, detail="Acesso administrativo negado.")
        data = bq.list_users()
        return {"users": data}

    elif action == "upsert":
        # Chamado no login — cria usuário se não existe, atualiza last_login se existe
        result = bq.upsert_user(
            google_sub=body.get("google_sub", ""),
            email=body.get("email", ""),
            nome=body.get("nome"),
            avatar_url=body.get("avatar_url"),
        )
        return {"success": True, **result}

    elif action == "update_role":
        # Só admin pode mudar roles
        caller_email = body.get("user_email", "")
        try:
            caller_is_admin = bq.is_admin(caller_email)
        except Exception:
            caller_is_admin = caller_email in settings.security.admin_emails
        if not caller_is_admin:
            raise HTTPException(status_code=403, detail="Acesso administrativo negado.")
        target_email = body.get("target_email", "")
        new_role = body.get("role", "user")
        if not target_email:
            raise HTTPException(status_code=400, detail="target_email obrigatório.")
        bq.update_user_role(target_email, new_role)
        return {"success": True, "email": target_email, "role": new_role}

    elif action == "check":
        # Verifica se email é admin — usado pelo frontend no login
        email = body.get("email", "")
        user = bq.get_user_by_email(email)
        if user:
            return {"exists": True, "role": user.get("role", "user"), "nome": user.get("nome", "")}
        return {"exists": False, "role": "user"}

    else:
        raise HTTPException(status_code=400, detail=f"Action '{action}' não suportada. Use: list, upsert, update_role, check.")


# ============================================================================
# POST /search-entities — Autocomplete de entidades (veículo, programa, praça)
# ============================================================================

class EntitySearchRequest(pydantic.BaseModel):
    query: str
    entity_type: str = "all"  # veiculo, programa, praca, plano, all

@app.post("/search-entities")
async def search_entities(request: EntitySearchRequest):
    """Busca entidades para autocomplete no input do chat."""
    from app.services.bq_service import get_bq_service
    bq = get_bq_service()
    q = request.query.strip().upper()
    if len(q) < 2:
        return {"results": []}

    results = []
    data_table = f"`{settings.bq.project_id}.{settings.bq.dataset_media}.pi01`"

    type_map = {
        "veiculo": ("veiculo", "Veículo"),
        "programa": ("programa", "Programa"),
        "praca": ("praca", "Praça"),
        "plano": ("plano_midia", "Plano"),
    }

    types_to_search = type_map if request.entity_type == "all" else {request.entity_type: type_map.get(request.entity_type, (request.entity_type, request.entity_type))}

    try:
        data_client = bigquery.Client(project=settings.bq.project_id)
        for etype, (col, label) in types_to_search.items():
            sql = f"""
                SELECT DISTINCT UPPER({col}) as name
                FROM {data_table}
                WHERE {col} IS NOT NULL AND UPPER({col}) LIKE @q
                ORDER BY name
                LIMIT 10
            """
            job_config = bigquery.QueryJobConfig(
                query_parameters=[bigquery.ScalarQueryParameter("q", "STRING", f"%{q}%")]
            )
            result = data_client.query(sql, job_config=job_config, timeout=10)
            for row in result.result(timeout=10):
                if row["name"]:
                    results.append({"name": row["name"], "type": etype, "label": label})
    except Exception as e:
        logger.warning("Falha em search-entities: %s", e)

    return {"results": results[:30]}


# ============================================================================
# GET /list-clients — Lista clientes disponíveis para o selector do front
# ============================================================================

@app.post("/list-clients")
async def list_clients():
    """Retorna lista de clientes únicos da base de dados.

    Usado pelo front para popular o selector de clientes.
    """
    from app.services.bq_service import get_bq_service
    bq = get_bq_service()
    clients = bq.list_clients()
    return {"clients": clients}


# ============================================================================
# POST /tts — Substitui o Webhook TTS do n8n (4 nodes)
# ============================================================================

@app.post("/tts", response_model=TTSResponse)
async def tts(request: TTSRequest):
    """Converte texto em audio. Google Cloud TTS (pt-BR nativo) com fallback OpenAI."""
    # Limitar texto a 4000 chars para evitar timeout
    text = request.text[:4000] if request.text else ""
    if not text.strip():
        raise HTTPException(status_code=400, detail="Texto vazio para TTS.")

    audio_b64 = None

    # Google Cloud TTS (pt-BR Neural2 nativo)
    if settings.tts.provider == "google":
        try:
            from google.cloud import texttospeech
            tts_client = texttospeech.TextToSpeechClient()
            synthesis_input = texttospeech.SynthesisInput(text=text)
            voice_params = texttospeech.VoiceSelectionParams(
                language_code=settings.tts.google_language,
                name=settings.tts.google_voice,
                ssml_gender=texttospeech.SsmlVoiceGender.MALE,
            )
            audio_config = texttospeech.AudioConfig(
                audio_encoding=texttospeech.AudioEncoding.MP3,
                speaking_rate=settings.tts.google_speaking_rate,
                pitch=0.0,
            )
            tts_response = tts_client.synthesize_speech(
                input=synthesis_input,
                voice=voice_params,
                audio_config=audio_config,
            )
            audio_b64 = base64.b64encode(tts_response.audio_content).decode("utf-8")
            logger.info("TTS Google Cloud gerado com sucesso (%d chars)", len(text))
        except Exception as google_err:
            logger.warning("Google TTS falhou, tentando OpenAI: %s", google_err)

    # Fallback OpenAI
    if audio_b64 is None:
        if not settings.tts.openai_api_key:
            raise HTTPException(status_code=503, detail="Nenhum provider TTS disponivel.")
        try:
            import openai
            client = openai.AsyncOpenAI(api_key=settings.tts.openai_api_key)
            response = await client.audio.speech.create(
                model=settings.tts.model,
                voice=settings.tts.voice,
                input=text,
            )
            audio_b64 = base64.b64encode(response.content).decode("utf-8")
            logger.info("TTS OpenAI gerado com sucesso (%d chars)", len(text))
        except Exception as e:
            logger.error("Erro no TTS: %s", e, exc_info=True)
            raise HTTPException(status_code=500, detail=f"Erro ao gerar audio: {e}")

    return TTSResponse(audio=audio_b64)


# ============================================================================
# POST /export — Substitui o Webhook Export Sheets do n8n (3 nodes)
# ============================================================================

@app.post("/export")
async def export_data(request: ExportRequest):
    """Exporta dados para Google Sheets, CSV ou XLSX.

    Substitui: Webhook Export → Prepare Data → Respond.
    format="sheets": cria Google Sheets nativo e compartilha com user_email.
    format="xlsx": retorna XLSX em base64.
    format="csv": retorna CSV em base64.
    """
    import io
    import csv
    import base64

    data = request.data
    if not data:
        return {"status": "error", "message": "Nenhum dado para exportar."}

    title = request.title or "athena_export"
    fmt = getattr(request, "format", "csv") or "csv"

    # ── Google Sheets nativo ──
    if fmt == "sheets":
        try:
            import gspread
            import google.auth
            from google.oauth2.credentials import Credentials as UserCredentials

            # Preferir token do usuario (cria no Drive DELE, nao da SA)
            if request.google_access_token:
                user_creds = UserCredentials(
                    token=request.google_access_token,
                    scopes=[
                        "https://www.googleapis.com/auth/spreadsheets",
                        "https://www.googleapis.com/auth/drive",
                    ],
                )
                gc = gspread.authorize(user_creds)
            else:
                # Fallback: ADC do Cloud Run (service account)
                creds, project = google.auth.default(
                    scopes=[
                        "https://www.googleapis.com/auth/spreadsheets",
                        "https://www.googleapis.com/auth/drive",
                    ]
                )
                gc = gspread.authorize(creds)

            # Criar planilha
            sh = gc.create(f"Athena — {title}")

            # Preparar dados
            ws = sh.sheet1
            ws.update_title(title[:100])

            if data and isinstance(data[0], dict):
                headers = list(data[0].keys())
                rows = [headers] + [[str(row.get(h, "")) for h in headers] for row in data]
            else:
                rows = [[str(c) for c in (row if isinstance(row, list) else [row])] for row in data]

            # Escrever tudo de uma vez (batch)
            ws.update(range_name="A1", values=rows)

            # Estilizar header (negrito + fundo vermelho)
            try:
                ws.format("1", {
                    "backgroundColor": {"red": 0.77, "green": 0.12, "blue": 0.12},
                    "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                })
            except Exception:
                pass  # Formatacao e cosmetic, nao bloqueia

            # Compartilhar com o usuario
            if request.user_email:
                try:
                    sh.share(request.user_email, perm_type="user", role="writer", notify=False)
                except Exception as e:
                    logger.warning(f"Falha ao compartilhar com {request.user_email}: {e}")
                    # Fallback: compartilhar com qualquer um que tenha o link
                    try:
                        sh.share("", perm_type="anyone", role="reader")
                    except Exception:
                        pass

            return {
                "status": "ok",
                "format": "sheets",
                "url": sh.url,
                "title": sh.title,
                "rows": len(data),
            }
        except ImportError:
            logger.error("gspread nao instalado, fallback para XLSX")
            fmt = "xlsx"
        except Exception as e:
            logger.error(f"Erro ao criar Google Sheet: {e}")
            return {"status": "error", "message": f"Erro ao criar planilha: {str(e)}"}

    if fmt == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title[:31]  # Excel limite 31 chars

            # Header
            if data and isinstance(data[0], dict):
                headers = list(data[0].keys())
                ws.append(headers)
                for row in data:
                    ws.append([row.get(h, "") for h in headers])
            else:
                for row in data:
                    ws.append(row if isinstance(row, list) else [str(row)])

            # Estilizar header
            from openpyxl.styles import Font, PatternFill
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C41E1E", end_color="C41E1E", fill_type="solid")

            # Auto-width
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

            buf = io.BytesIO()
            wb.save(buf)
            b64 = base64.b64encode(buf.getvalue()).decode()
            return {
                "status": "ok",
                "format": "xlsx",
                "filename": f"{title}.xlsx",
                "content_base64": b64,
                "rows": len(data),
            }
        except ImportError:
            logger.warning("openpyxl não instalado, fallback para CSV")
            fmt = "csv"

    # CSV fallback
    buf = io.StringIO()
    if data and isinstance(data[0], dict):
        writer = csv.DictWriter(buf, fieldnames=list(data[0].keys()))
        writer.writeheader()
        writer.writerows(data)
    else:
        writer = csv.writer(buf)
        for row in data:
            writer.writerow(row if isinstance(row, list) else [str(row)])

    b64 = base64.b64encode(buf.getvalue().encode("utf-8-sig")).decode()
    return {
        "status": "ok",
        "format": "csv",
        "filename": f"{title}.csv",
        "content_base64": b64,
        "rows": len(data),
    }


# ============================================================================
# GET/POST /settings/domains — Gerenciar domínios de e-mail permitidos
# ============================================================================

@app.get("/settings/domains")
async def get_allowed_domains():
    """Retorna lista de domínios de e-mail permitidos para login."""
    from app.services.bq_service import get_bq_service
    bq = get_bq_service()
    try:
        sql = f"""
            SELECT domain FROM `{settings.bq.project_persistence}.{settings.bq.dataset_persistence}.athena_settings`
            WHERE setting_key = 'allowed_domain'
            ORDER BY domain
        """
        result = bq._client_persistence.query(sql, timeout=10)
        domains = [row["domain"] for row in result.result(timeout=10)]
        if not domains:
            # Fallback: domínios padrão
            domains = ["grupoom.com.br", "opusmultipla.com.br"]
        return {"domains": domains}
    except Exception:
        return {"domains": ["grupoom.com.br", "opusmultipla.com.br"]}


class DomainRequest(pydantic.BaseModel):
    domain: str


@app.post("/settings/domains/add")
async def add_allowed_domain(request: DomainRequest):
    """Adiciona um domínio de e-mail permitido."""
    domain = request.domain.strip().lower()
    if not domain or "." not in domain:
        raise HTTPException(status_code=400, detail="Domínio inválido")

    from app.services.bq_service import get_bq_service
    bq = get_bq_service()
    try:
        # Cria tabela se não existe
        bq._client_persistence.query(f"""
            CREATE TABLE IF NOT EXISTS `{settings.bq.project_persistence}.{settings.bq.dataset_persistence}.athena_settings` (
                setting_key STRING,
                domain STRING,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """, timeout=15).result(timeout=15)

        # Insere domínio (evita duplicata)
        bq._client_persistence.query(f"""
            INSERT INTO `{settings.bq.project_persistence}.{settings.bq.dataset_persistence}.athena_settings`
            (setting_key, domain)
            SELECT 'allowed_domain', @domain
            WHERE NOT EXISTS (
                SELECT 1 FROM `{settings.bq.project_persistence}.{settings.bq.dataset_persistence}.athena_settings`
                WHERE setting_key = 'allowed_domain' AND domain = @domain
            )
        """, job_config=bigquery.QueryJobConfig(
            query_parameters=[bigquery.ScalarQueryParameter("domain", "STRING", domain)]
        ), timeout=15).result(timeout=15)

        return {"ok": True, "domain": domain}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/settings/domains/remove")
async def remove_allowed_domain(request: DomainRequest):
    """Remove um domínio de e-mail permitido."""
    domain = request.domain.strip().lower()
    from app.services.bq_service import get_bq_service
    bq = get_bq_service()
    try:
        bq._client_persistence.query(f"""
            DELETE FROM `{settings.bq.project_persistence}.{settings.bq.dataset_persistence}.athena_settings`
            WHERE setting_key = 'allowed_domain' AND domain = @domain
        """, job_config=bigquery.QueryJobConfig(
            query_parameters=[bigquery.ScalarQueryParameter("domain", "STRING", domain)]
        ), timeout=15).result(timeout=15)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# GET/POST /settings/synonyms — Dicionário de sinônimos
# ============================================================================

@app.get("/settings/synonyms")
async def get_synonyms():
    """Retorna dicionário de sinônimos armazenado no BigQuery."""
    try:
        data_client = bigquery.Client(project=settings.bq.project_persistence)
        table = f"`{settings.bq.project_persistence}.{settings.bq.dataset_persistence}.athena_settings`"
        sql = f"""
            SELECT domain as term_from, setting_value as term_to
            FROM {table}
            WHERE setting_key = 'synonym'
            ORDER BY domain
        """
        result = data_client.query(sql, timeout=15).result(timeout=15)
        synonyms = [{"from": row["term_from"], "to": row["term_to"]} for row in result]
        return {"synonyms": synonyms}
    except Exception:
        return {"synonyms": []}


@app.post("/settings/synonyms/add")
async def add_synonym(request: Request):
    """Adiciona sinônimo ao dicionário."""
    body = await request.json()
    term_from = body.get("term_from", "").strip().lower()
    term_to = body.get("term_to", "").strip()
    if not term_from or not term_to:
        raise HTTPException(status_code=400, detail="term_from e term_to obrigatórios")
    try:
        data_client = bigquery.Client(project=settings.bq.project_persistence)
        table = f"`{settings.bq.project_persistence}.{settings.bq.dataset_persistence}.athena_settings`"
        sql = f"""
            INSERT INTO {table} (setting_key, domain, setting_value, created_at)
            VALUES ('synonym', @term_from, @term_to, CURRENT_TIMESTAMP())
        """
        data_client.query(sql, job_config=bigquery.QueryJobConfig(
            query_parameters=[
                bigquery.ScalarQueryParameter("term_from", "STRING", term_from),
                bigquery.ScalarQueryParameter("term_to", "STRING", term_to),
            ]
        ), timeout=15).result(timeout=15)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/settings/synonyms/remove")
async def remove_synonym(request: Request):
    """Remove sinônimo do dicionário."""
    body = await request.json()
    term_from = body.get("term_from", "").strip().lower()
    if not term_from:
        raise HTTPException(status_code=400, detail="term_from obrigatório")
    try:
        data_client = bigquery.Client(project=settings.bq.project_persistence)
        table = f"`{settings.bq.project_persistence}.{settings.bq.dataset_persistence}.athena_settings`"
        sql = f"""
            DELETE FROM {table}
            WHERE setting_key = 'synonym' AND domain = @term_from
        """
        data_client.query(sql, job_config=bigquery.QueryJobConfig(
            query_parameters=[bigquery.ScalarQueryParameter("term_from", "STRING", term_from)]
        ), timeout=15).result(timeout=15)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Upload de Documentos (PDF/Excel) — Feature Caroline/Phillipe
# ============================================================================

@app.post("/upload")
async def upload_document(file: UploadFile = File(...)):
    """Recebe PDF ou Excel, extrai texto e tabelas, retorna conteudo estruturado.

    Usado para: propostas de veiculos, planilhas de entrega, tabelas de custo.
    Limite: 10MB.
    Formatos: PDF, XLSX, XLS, CSV.
    """
    # Auth ja validada pelo middleware global (Bearer token)

    if not file.filename:
        raise HTTPException(status_code=400, detail="Nenhum arquivo enviado.")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    allowed = {"pdf", "xlsx", "xls", "csv"}
    if ext not in allowed:
        raise HTTPException(
            status_code=400,
            detail=f"Formato .{ext} não suportado. Use: {', '.join(allowed)}",
        )

    contents = await file.read()
    max_size = 10 * 1024 * 1024  # 10MB
    if len(contents) > max_size:
        raise HTTPException(status_code=400, detail="Arquivo excede 10MB.")

    try:
        if ext == "pdf":
            extracted = _extract_pdf(contents, file.filename)
        elif ext in ("xlsx", "xls"):
            extracted = _extract_excel(contents, file.filename)
        elif ext == "csv":
            extracted = _extract_csv(contents, file.filename)
        else:
            raise HTTPException(status_code=400, detail="Formato não suportado.")

        logger.info(
            "Upload processado: %s (%s, %d bytes, %d chars extraidos)",
            file.filename, ext, len(contents), len(extracted.get("text", "")),
        )
        return extracted

    except HTTPException:
        raise
    except Exception as e:
        logger.error("Erro ao processar upload %s: %s", file.filename, e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")


def _extract_pdf(contents: bytes, filename: str) -> dict:
    """Extrai texto e tabelas de PDF usando pdfplumber."""
    import io
    import pdfplumber

    text_parts = []
    tables_found = []

    with pdfplumber.open(io.BytesIO(contents)) as pdf:
        for i, page in enumerate(pdf.pages):
            # Texto da pagina
            page_text = page.extract_text()
            if page_text:
                text_parts.append(f"--- Página {i + 1} ---\n{page_text}")

            # Tabelas da pagina
            page_tables = page.extract_tables()
            for ti, table in enumerate(page_tables):
                if table and len(table) > 1:
                    # Primeira linha = headers
                    headers = [str(c or "").strip() for c in table[0]]
                    rows = []
                    for row in table[1:]:
                        rows.append([str(c or "").strip() for c in row])
                    tables_found.append({
                        "page": i + 1,
                        "index": ti,
                        "headers": headers,
                        "rows": rows,
                    })

    full_text = "\n\n".join(text_parts)

    # Limita texto a 30k chars para nao estourar contexto do LLM
    if len(full_text) > 30000:
        full_text = full_text[:30000] + "\n\n[... texto truncado, arquivo muito grande ...]"

    return {
        "filename": filename,
        "type": "pdf",
        "pages": len(text_parts),
        "text": full_text,
        "tables": tables_found,
        "tables_count": len(tables_found),
    }


def _extract_excel(contents: bytes, filename: str) -> dict:
    """Extrai dados de Excel usando openpyxl."""
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    sheets_data = []
    text_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c if c is not None else "").strip() for c in row])

        if not rows:
            continue

        # Primeira linha = headers
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        # Limita a 500 linhas por aba
        if len(data_rows) > 500:
            data_rows = data_rows[:500]

        sheets_data.append({
            "sheet": sheet_name,
            "headers": headers,
            "rows": data_rows,
            "total_rows": len(rows) - 1,
        })

        # Texto tabular para contexto do LLM
        text_parts.append(f"--- Aba: {sheet_name} ({len(data_rows)} linhas) ---")
        text_parts.append(" | ".join(headers))
        for dr in data_rows[:100]:  # max 100 linhas no texto
            text_parts.append(" | ".join(dr))

    wb.close()
    full_text = "\n".join(text_parts)

    if len(full_text) > 30000:
        full_text = full_text[:30000] + "\n\n[... texto truncado ...]"

    return {
        "filename": filename,
        "type": "excel",
        "sheets": sheets_data,
        "sheets_count": len(sheets_data),
        "text": full_text,
    }


def _extract_csv(contents: bytes, filename: str) -> dict:
    """Extrai dados de CSV."""
    import csv
    import io

    # Tenta detectar encoding
    try:
        text = contents.decode("utf-8")
    except UnicodeDecodeError:
        text = contents.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    if not rows:
        return {"filename": filename, "type": "csv", "text": "", "headers": [], "rows": []}

    headers = rows[0]
    data_rows = rows[1:500]  # max 500

    text_parts = [" | ".join(headers)]
    for r in data_rows[:100]:
        text_parts.append(" | ".join(r))

    return {
        "filename": filename,
        "type": "csv",
        "headers": headers,
        "rows": data_rows,
        "total_rows": len(rows) - 1,
        "text": "\n".join(text_parts),
    }


# ============================================================================
# Entrypoint
# ============================================================================

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "app.main:app",
        host=settings.host,
        port=settings.port,
        reload=settings.debug,
    )
